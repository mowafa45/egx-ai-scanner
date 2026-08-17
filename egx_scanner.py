"""
EGX AI Scanner v2 - Higher Frequency Breakout + Pullback Strategy
==================================================================
يفحص أسهم البورصة المصرية ويبعت الفرص على تليجرام كـ Excel
"""

import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime
import asyncio, logging, io, warnings
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import telegram
from config import CONFIG

warnings.filterwarnings("ignore")
logging.getLogger("yfinance").setLevel(logging.CRITICAL)
logging.getLogger("peewee").setLevel(logging.CRITICAL)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("egx_scanner.log", encoding="utf-8")
    ]
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# EGX TICKERS  (رموز Yahoo Finance الصحيحة)
# ─────────────────────────────────────────────
EGX_TICKERS = [
    # بنوك
    "COMI.CA", "QNBE.CA", "ADIB.CA", "EGBE.CA", "CIEB.CA",
    "HDBK.CA", "MCQE.CA", "FAIT.CA", "CANA.CA",
    # اتصالات وتكنولوجيا
    "ETEL.CA", "SWDY.CA", "RAYA.CA", "FWRY.CA", "EFIH.CA",
    # عقارات
    "TMGH.CA", "EMFD.CA", "PHDC.CA", "OCDI.CA", "ORHD.CA",
    "MASR.CA", "MNHD.CA",
    # صناعة وكيماويات
    "HRHO.CA", "ABUK.CA", "MFPC.CA", "ARCC.CA", "EGAL.CA",
    "IRON.CA", "CLHO.CA", "EGCH.CA", "FERC.CA", "SKPC.CA",
    # غذاء ومشروبات
    "JUFO.CA", "SUGR.CA", "POUL.CA", "DOMT.CA", "EFID.CA",
    # طاقة
    "AMOC.CA",
    # صحة
    "ISPH.CA",
    # مالية
    "EGTS.CA", "ALCN.CA", "BTFH.CA", "GBCO.CA", "VLMR.CA",
    # بنية تحتية
    "ORAS.CA", "AMER.CA", "CIRA.CA", "EAST.CA", "SCTS.CA",
    # متنوع
    "MENA.CA", "ALUM.CA", "GPPL.CA",
]
EGX_TICKERS = list(dict.fromkeys(EGX_TICKERS))


# ─────────────────────────────────────────────
# INDICATORS
# ─────────────────────────────────────────────
def sma(s, n): return s.rolling(n).mean()

def rsi(s, n=14):
    d = s.diff()
    g = d.clip(lower=0).rolling(n).mean()
    lo = (-d.clip(upper=0)).rolling(n).mean()
    return 100 - 100 / (1 + g / lo.replace(0, np.nan))

def atr(h, lo, c, n=14):
    tr = pd.concat([h - lo, (h - c.shift()).abs(), (lo - c.shift()).abs()], axis=1).max(axis=1)
    return tr.rolling(n).mean()

def adx(h, lo, c, n=14):
    pdm = h.diff().clip(lower=0)
    ndm = (-lo.diff()).clip(lower=0)
    pdm[pdm < ndm] = 0
    ndm[ndm < pdm] = 0
    a = atr(h, lo, c, n)
    pdi = 100 * pdm.rolling(n).mean() / a
    ndi = 100 * ndm.rolling(n).mean() / a
    dx = 100 * (pdi - ndi).abs() / (pdi + ndi).replace(0, np.nan)
    return dx.rolling(n).mean()

def is_rising(s, n):
    """ta.rising(close, n) — كل شمعة أعلى من السابقة لـ n شمعات"""
    result = pd.Series(True, index=s.index)
    for i in range(1, n + 1):
        result = result & (s > s.shift(i))
    return result


# ─────────────────────────────────────────────
# DATA FETCH  (مع retry)
# ─────────────────────────────────────────────
def fetch(ticker: str) -> pd.DataFrame | None:
    import time
    for attempt in range(3):
        try:
            t = yf.Ticker(ticker)
            df = t.history(period="2y", interval="1d", auto_adjust=True, timeout=20)

            if df is None or df.empty or len(df) < 50:
                return None

            keep = ["Open", "High", "Low", "Close", "Volume"]
            df = df[[c for c in keep if c in df.columns]].copy()

            if not all(c in df.columns for c in keep):
                return None

            # إصلاح Volume: نضمن إنه int64 حقيقي (مش float أو NaN)
            df["Volume"] = df["Volume"].fillna(0).astype("int64")

            # لو آخر شمعة حجمها صفر (يوم لسه ما اتفتح) نشيلها
            if df["Volume"].iloc[-1] == 0:
                df = df.iloc[:-1]

            if df.empty or len(df) < 50 or df["Volume"].sum() == 0:
                return None

            return df

        except Exception:
            pass
        if attempt < 2:
            time.sleep(1.5)
    return None


# ─────────────────────────────────────────────
# STRATEGY CHECK  (مطابق لـ Pine Script v2)
# شرطان للدخول: Breakout أو Pullback
# ─────────────────────────────────────────────
def check_strategy(ticker: str, cfg: dict) -> dict | None:
    df = fetch(ticker)
    if df is None or len(df) < cfg["sma_length"] + 10:
        return None

    c  = df["Close"].squeeze()
    h  = df["High"].squeeze()
    lo = df["Low"].squeeze()
    v  = df["Volume"].squeeze()
    o  = df["Open"].squeeze()

    s200      = sma(c, cfg["sma_length"])        # SMA200 — فلتر الاتجاه
    r14       = rsi(c, cfg["rsi_length"])
    a14       = atr(h, lo, c, cfg["atr_length"])
    vs        = sma(v, cfg["vol_lookback"])
    adx14     = adx(h, lo, c, cfg["adx_length"])
    ph        = h.rolling(cfg["breakout_lookback"]).max().shift(1)
    rising    = is_rising(c, cfg["pullback_sensitivity"])

    cv, ov, vv = float(c.iloc[-1]), float(o.iloc[-1]), float(v.iloc[-1])
    s200v      = float(s200.iloc[-1])
    rv         = float(r14.iloc[-1])
    av         = float(a14.iloc[-1])
    vsv        = float(vs.iloc[-1])
    adxv       = float(adx14.iloc[-1])
    phv        = float(ph.iloc[-1])
    lv         = float(lo.iloc[-1])
    rising_now = bool(rising.iloc[-1])

    if any(np.isnan([cv, ov, vv, s200v, rv, av, phv])):
        return None
    if vsv == 0:
        return None

    # ── الشروط المشتركة (مطابق للـ Pine Script) ──
    trend_bullish = cv > s200v                          # close > SMA200
    strong_volume = vv > vsv * cfg["vol_multiplier"]
    healthy_rsi   = cfg["rsi_lower"] <= rv < cfg["rsi_upper"]
    strong_trend  = adxv >= cfg["adx_threshold"]

    if not (trend_bullish and strong_volume and healthy_rsi and strong_trend):
        return None

    # ── نوع الدخول ──
    breakout = cv > phv and cv > ov                     # كسر أعلى سابق
    pullback = (lv <= s200v * 1.015                     # لمس SMA200
                and cv > s200v
                and rising_now)

    use_pullback = cfg.get("use_pullback_mode", True)

    if breakout:
        signal_type = "Breakout"
    elif use_pullback and pullback:
        signal_type = "Pullback"
    else:
        return None

    sl  = cv - av * cfg["atr_mult"]
    tp  = cv + av * cfg["tp_atr_mult"]
    rr  = (tp - cv) / max(cv - sl, 0.0001)
    pv  = float(c.iloc[-2]) if len(c) > 1 else cv

    return {
        "ticker":       ticker.replace(".CA", ""),
        "signal_type":  signal_type,
        "close":        round(cv, 2),
        "change_pct":   round((cv - pv) / pv * 100, 2),
        "sma200":       round(s200v, 2),
        "rsi":          round(rv, 1),
        "adx":          round(adxv, 1),
        "atr":          round(av, 2),
        "vol_ratio":    round(vv / vsv, 2),
        "prev_high":    round(phv, 2),
        "stop_loss":    round(sl, 2),
        "take_profit":  round(tp, 2),
        "risk_reward":  round(rr, 2),
        "scan_time":    datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


# ─────────────────────────────────────────────
# DIAGNOSE: ليه السهم مش مؤهل (للـ log)
# ─────────────────────────────────────────────
def diagnose(ticker: str, cfg: dict) -> str:
    df = fetch(ticker)
    if df is None:
        return "❌ لا توجد بيانات"
    c  = df["Close"].squeeze(); h = df["High"].squeeze()
    lo = df["Low"].squeeze(); v = df["Volume"].squeeze(); o = df["Open"].squeeze()
    cv, ov, vv = float(c.iloc[-1]), float(o.iloc[-1]), float(v.iloc[-1])
    lv     = float(lo.iloc[-1])
    s200v  = float(sma(c, cfg["sma_length"]).iloc[-1])
    rv     = float(rsi(c, cfg["rsi_length"]).iloc[-1])
    vsv    = float(sma(v, cfg["vol_lookback"]).iloc[-1])
    adxv   = float(adx(h, lo, c, cfg["adx_length"]).iloc[-1])
    phv    = float(h.rolling(cfg["breakout_lookback"]).max().shift(1).iloc[-1])
    rising_now = bool(is_rising(c, cfg["pullback_sensitivity"]).iloc[-1])

    reasons = []
    if not cv > s200v:
        reasons.append(f"تحت SMA200 ({cv:.1f}<{s200v:.1f})")
    if not (cfg["rsi_lower"] <= rv < cfg["rsi_upper"]):
        reasons.append(f"RSI={rv:.1f}")
    if vsv > 0 and not vv > vsv * cfg["vol_multiplier"]:
        reasons.append(f"حجم {vv/vsv:.1f}x")
    elif vsv == 0:
        reasons.append("حجم: متوسط=0")
    if not adxv >= cfg["adx_threshold"]:
        reasons.append(f"ADX={adxv:.1f}")
    breakout = cv > phv and cv > ov
    pullback = lv <= s200v * 1.015 and cv > s200v and rising_now
    if not breakout and not (cfg.get("use_pullback_mode", True) and pullback):
        if not breakout:
            reasons.append(f"لم يكسر {phv:.1f}")
        if cfg.get("use_pullback_mode", True) and not pullback:
            reasons.append("لا pullback")

    return " | ".join(reasons) if reasons else "✅"


# ─────────────────────────────────────────────
# SCORE: درجة قرب السهم من تحقيق الشروط (0-100)
# ─────────────────────────────────────────────
def score_ticker(ticker: str, cfg: dict) -> dict | None:
    """
    يحسب درجة قرب السهم من الشروط لو مش مؤهل كاملاً.
    بيرجع dict فيه الدرجة وإيه الناقص بالأرقام.
    """
    df = fetch(ticker)
    if df is None or len(df) < cfg["sma_length"] + 10:
        return None

    c  = df["Close"].squeeze(); h = df["High"].squeeze()
    lo = df["Low"].squeeze();   v = df["Volume"].squeeze(); o = df["Open"].squeeze()

    cv, ov, vv = float(c.iloc[-1]), float(o.iloc[-1]), float(v.iloc[-1])
    lv         = float(lo.iloc[-1])
    s200v      = float(sma(c, cfg["sma_length"]).iloc[-1])
    rv         = float(rsi(c, cfg["rsi_length"]).iloc[-1])
    av         = float(atr(h, lo, c, cfg["atr_length"]).iloc[-1])
    vsv        = float(sma(v, cfg["vol_lookback"]).iloc[-1])
    adxv       = float(adx(h, lo, c, cfg["adx_length"]).iloc[-1])
    phv        = float(h.rolling(cfg["breakout_lookback"]).max().shift(1).iloc[-1])
    rising_now = bool(is_rising(c, cfg["pullback_sensitivity"]).iloc[-1])
    pv         = float(c.iloc[-2]) if len(c) > 1 else cv

    if any(np.isnan([cv, ov, vv, s200v, rv, av, phv])) or vsv == 0:
        return None

    score = 0
    missing = []

    # ── الشروط المشتركة (كل شرط = 20 نقطة) ──
    # 1. Trend: close > SMA200
    if cv > s200v:
        score += 20
    else:
        gap_pct = (s200v - cv) / s200v * 100
        missing.append(f"تحت SMA200 بـ {gap_pct:.1f}%")

    # 2. RSI
    rsi_mid = (cfg["rsi_lower"] + cfg["rsi_upper"]) / 2
    if cfg["rsi_lower"] <= rv < cfg["rsi_upper"]:
        score += 20
    else:
        if rv < cfg["rsi_lower"]:
            missing.append(f"RSI ضعيف ({rv:.1f} < {cfg['rsi_lower']})")
        else:
            missing.append(f"RSI مشبع ({rv:.1f} > {cfg['rsi_upper']})")

    # 3. Volume
    vol_ratio = vv / vsv
    if vol_ratio >= cfg["vol_multiplier"]:
        score += 20
    else:
        missing.append(f"حجم {vol_ratio:.1f}x (محتاج {cfg['vol_multiplier']}x)")

    # 4. ADX
    if adxv >= cfg["adx_threshold"]:
        score += 20
    else:
        missing.append(f"ADX {adxv:.1f} (محتاج {cfg['adx_threshold']})")

    # 5. Entry type (20 نقطة)
    breakout = cv > phv and cv > ov
    pullback = lv <= s200v * 1.015 and cv > s200v and rising_now
    if breakout:
        score += 20
        entry_status = f"كسر {phv:.2f} ✅"
    elif cfg.get("use_pullback_mode", True) and pullback:
        score += 20
        entry_status = "Pullback ✅"
    else:
        dist_to_break = (phv - cv) / cv * 100
        entry_status = f"يحتاج {dist_to_break:.1f}% للكسر"
        missing.append(entry_status)

    sl = cv - av * cfg["atr_mult"]
    tp = cv + av * cfg["tp_atr_mult"]
    rr = (tp - cv) / max(cv - sl, 0.0001)

    return {
        "ticker":       ticker.replace(".CA", ""),
        "score":        score,
        "missing":      " | ".join(missing) if missing else "✅ جاهز",
        "close":        round(cv, 2),
        "change_pct":   round((cv - pv) / pv * 100, 2),
        "sma200":       round(s200v, 2),
        "rsi":          round(rv, 1),
        "adx":          round(adxv, 1),
        "vol_ratio":    round(vol_ratio, 2),
        "prev_high":    round(phv, 2),
        "stop_loss":    round(sl, 2),
        "take_profit":  round(tp, 2),
        "risk_reward":  round(rr, 2),
        "entry_status": entry_status,
        "scan_time":    datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


# ─────────────────────────────────────────────
# SCAN
# ─────────────────────────────────────────────
def run_scan(tickers: list, cfg: dict) -> tuple[list, list, dict]:
    signals   = []
    candidates = []   # أقرب الأسهم لتحقيق الشروط
    stats     = {"total": len(tickers), "data_ok": 0, "no_data": 0}

    for idx, t in enumerate(tickers, 1):
        log.info(f"[{idx:>3}/{len(tickers)}] {t}")
        df = fetch(t)

        if df is None:
            stats["no_data"] += 1
            log.warning(f"  ⚠️  لا توجد بيانات")
            continue

        stats["data_ok"] += 1
        result = check_strategy(t, cfg)

        if result:
            signals.append(result)
            sig_icon = "🔺" if result["signal_type"] == "Breakout" else "🔄"
            log.info(f"  ✅ {sig_icon} {result['signal_type']}! {result['ticker']} | "
                     f"Close={result['close']} | RSI={result['rsi']} | "
                     f"ADX={result['adx']} | R:R={result['risk_reward']}x")
        else:
            # حساب درجة القرب لو مش مؤهل
            scored = score_ticker(t, cfg)
            if scored and scored["score"] >= 40:   # على الأقل حقق شرطين
                candidates.append(scored)
                log.info(f"  🟡 {t.replace('.CA','')} [{scored['score']}/100]: {scored['missing']}")
            else:
                why = diagnose(t, cfg)
                log.info(f"  ➖ {t.replace('.CA','')}: {why}")

    # أعلى 3 مرشحين بالدرجة
    candidates.sort(key=lambda x: x["score"], reverse=True)
    top3 = candidates[:3]

    return signals, top3, stats


# ─────────────────────────────────────────────
# EXCEL REPORT
# ─────────────────────────────────────────────
def build_excel(signals: list, stats: dict, near_miss: list = None) -> bytes:
    if near_miss is None: near_miss = []
    wb = Workbook()
    ws = wb.active
    ws.title = "EGX Signals"

    BG, GRN, GLD, WHT = "0D1117", "00C896", "FFD700", "FFFFFF"

    # Title
    ws.merge_cells("A1:N1")
    c1 = ws["A1"]
    c1.value = f"🔍  EGX Breakout + Pullback Scanner  |  {datetime.now().strftime('%Y-%m-%d  %H:%M')}"
    c1.font = Font(name="Calibri", bold=True, size=16, color=GLD)
    c1.fill = PatternFill("solid", fgColor=BG)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Stats
    ws.merge_cells("A2:N2")
    c2 = ws["A2"]
    c2.value = (f"فرص اليوم: {len(signals)}  |  فُحص: {stats['data_ok']}  |  "
                f"بدون بيانات: {stats['no_data']}  |  "
                f"الاستراتيجية: SMA200 + Breakout/Pullback + Volume + RSI + ADX")
    c2.font = Font(name="Calibri", size=10, color="8B949E")
    c2.fill = PatternFill("solid", fgColor=BG)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    headers = ["الكود", "النوع", "السعر", "التغير%", "SMA200", "RSI", "ADX",
               "نسبة الحجم", "وقف الخسارة", "الهدف", "R:R",
               "ATR", "أعلى سابق", "وقت الفحص"]
    widths  = [10, 11, 12, 10, 12, 8, 8, 11, 14, 14, 8, 10, 14, 18]

    for col, (hdr, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col, value=hdr)
        cell.font = Font(name="Calibri", bold=True, size=11, color=WHT)
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=GRN))
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 32

    if not signals:
        ws.merge_cells("A4:N4")
        nc = ws["A4"]
        nc.value = "لا توجد فرص تحقق شروط الاستراتيجية حالياً"
        nc.font = Font(name="Calibri", size=12, color="8B949E", italic=True)
        nc.fill = PatternFill("solid", fgColor=BG)
        nc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[4].height = 28
    else:
        for ri, sig in enumerate(signals, 4):
            bg_clr = "161B22" if ri % 2 == 0 else BG
            sig_type_label = "🔺 Breakout" if sig["signal_type"] == "Breakout" else "🔄 Pullback"
            row_data = [
                sig["ticker"], sig_type_label, sig["close"], sig["change_pct"] / 100,
                sig["sma200"], sig["rsi"], sig["adx"], sig["vol_ratio"],
                sig["stop_loss"], sig["take_profit"], sig["risk_reward"],
                sig["atr"], sig["prev_high"], sig["scan_time"]
            ]
            fmts = [None, None, "#,##0.00", "0.00%", "#,##0.00", "0.0", "0.0",
                    '0.00"x"', "#,##0.00", "#,##0.00", '0.00"x"',
                    "#,##0.00", "#,##0.00", None]

            for col, (val, fmt) in enumerate(zip(row_data, fmts), 1):
                cell = ws.cell(row=ri, column=col, value=val)
                cell.fill = PatternFill("solid", fgColor=bg_clr)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(bottom=Side(style="thin", color="1F2937"))
                if fmt:
                    cell.number_format = fmt

                # Colors (col1=ticker,col2=type,col3=price,col4=chg%,col5=sma200,col6=RSI,col7=ADX,col8=vol...)
                if col == 1:
                    cell.font = Font(name="Calibri", bold=True, size=12, color=GRN)
                elif col == 2:
                    clr = "00C896" if sig["signal_type"] == "Breakout" else "90EE90"
                    cell.font = Font(name="Calibri", bold=True, color=clr)
                elif col == 4:
                    clr = "00C896" if sig["change_pct"] >= 0 else "FF4C4C"
                    cell.font = Font(name="Calibri", bold=True, color=clr)
                elif col == 9:
                    cell.font = Font(name="Calibri", bold=True, color="FF4C4C")
                elif col == 10:
                    cell.font = Font(name="Calibri", bold=True, color="00C896")
                elif col == 11:
                    clr = "00C896" if sig["risk_reward"] >= 1.5 else "FFD700"
                    cell.font = Font(name="Calibri", bold=True, color=clr)
                elif col in [6, 7, 8]:
                    cell.font = Font(name="Calibri", color=GLD)
                elif col == 14:
                    cell.font = Font(name="Calibri", color="8B949E")
                else:
                    cell.font = Font(name="Calibri", color=WHT)

            ws.row_dimensions[ri].height = 24

    ws.freeze_panes = "A4"
    ws.sheet_properties.tabColor = "00C896"

    # ── Near-Miss sheet (أقرب 3 أسهم)
    ws_near = wb.create_sheet("🟡 قريب من الشروط")
    ws_near.merge_cells("A1:I1")
    nt = ws_near["A1"]
    nt.value = "🟡  أقرب الأسهم لتحقيق الشروط  —  تابعهم في الجلسات القادمة"
    nt.font = Font(name="Calibri", bold=True, size=14, color=GLD)
    nt.fill = PatternFill("solid", fgColor=BG)
    nt.alignment = Alignment(horizontal="center", vertical="center")
    ws_near.row_dimensions[1].height = 30

    near_hdrs = ["الكود", "الدرجة", "السعر", "RSI", "ADX", "نسبة الحجم", "أعلى سابق", "وضع الدخول", "الناقص"]
    near_wdts = [10, 9, 12, 8, 8, 12, 14, 18, 45]
    for col, (hdr, w) in enumerate(zip(near_hdrs, near_wdts), 1):
        cell = ws_near.cell(row=2, column=col, value=hdr)
        cell.font = Font(name="Calibri", bold=True, size=11, color=WHT)
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color="FFD700"))
        ws_near.column_dimensions[get_column_letter(col)].width = w
    ws_near.row_dimensions[2].height = 28

    if not near_miss:
        ws_near.merge_cells("A3:I3")
        nc = ws_near["A3"]
        nc.value = "لا توجد أسهم قريبة من الشروط حالياً"
        nc.font = Font(name="Calibri", size=11, color="8B949E", italic=True)
        nc.fill = PatternFill("solid", fgColor=BG)
        nc.alignment = Alignment(horizontal="center", vertical="center")
    else:
        for ri, nm in enumerate(near_miss, 3):
            bg_clr = "161B22" if ri % 2 == 0 else BG
            score_color = "00C896" if nm["score"] >= 80 else ("FFD700" if nm["score"] >= 60 else "FF9900")
            row_vals = [
                nm["ticker"], str(nm["score"]) + "/100",
                nm["close"], nm["rsi"], nm["adx"],
                str(nm["vol_ratio"]) + "x",
                nm["prev_high"], nm["entry_status"], nm["missing"]
            ]
            for col, val in enumerate(row_vals, 1):
                cell = ws_near.cell(row=ri, column=col, value=val)
                cell.fill = PatternFill("solid", fgColor=bg_clr)
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=(col == 9))
                cell.border = Border(bottom=Side(style="thin", color="1F2937"))
                if col == 1:
                    cell.font = Font(name="Calibri", bold=True, size=12, color=GRN)
                elif col == 2:
                    cell.font = Font(name="Calibri", bold=True, color=score_color)
                elif col == 9:
                    cell.font = Font(name="Calibri", size=10, color="FF9900")
                    cell.alignment = Alignment(horizontal="right", vertical="center",
                                               wrap_text=True)
                else:
                    cell.font = Font(name="Calibri", color=WHT)
            ws_near.row_dimensions[ri].height = 36
    ws_near.freeze_panes = "A3"
    ws_near.sheet_properties.tabColor = "FFD700"

    # Legend sheet
    ws2 = wb.create_sheet("📖 شرح")
    legend = [
        ("المؤشر", "الوصف", "القيمة"),
        ("SMA 200", "متوسط متحرك 200 يوم", f"السعر > SMA200"),
        ("RSI", "زخم الحركة", f"{CONFIG['rsi_lower']}–{CONFIG['rsi_upper']}"),
        ("ADX", "قوة الاتجاه", f"≥ {CONFIG['adx_threshold']}"),
        ("Breakout", "كسر الأعلى السابق", f"آخر {CONFIG['breakout_lookback']} يوم"),
        ("Pullback", "ارتداد من SMA200", "Low ≤ SMA×1.015 + صاعد"),
        ("Volume", "تأكيد الاختراق", f"> {CONFIG['vol_multiplier']}x المتوسط"),
        ("SL", "وقف الخسارة", f"سعر − {CONFIG['atr_mult']} × ATR"),
        ("TP", "الهدف", f"سعر + {CONFIG['tp_atr_mult']} × ATR"),
        ("R:R", "عائد/مخاطرة", "≥ 1.5 أفضل"),
    ]
    for r, row in enumerate(legend, 1):
        for col, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=col, value=val)
            if r == 1:
                cell.font = Font(bold=True, color=GLD, size=12)
                cell.fill = PatternFill("solid", fgColor="1F2937")
            else:
                cell.font = Font(color=WHT)
                cell.fill = PatternFill("solid", fgColor="161B22" if r % 2 == 0 else BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=Side(style="thin", color="1F2937"))
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────
# TELEGRAM
# ─────────────────────────────────────────────
async def send_telegram(signals: list, excel_bytes: bytes, stats: dict, cfg: dict, near_miss: list = None):
    if near_miss is None: near_miss = []
    token = cfg.get("telegram_token", "")
    if not token or token == "YOUR_BOT_TOKEN_HERE":
        log.warning("Telegram غير مُعدّ في config.py")
        return

    bot = telegram.Bot(token=token)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    lines = ["🟢 EGX Breakout Scanner" if signals else "🔍 EGX Breakout Scanner",
             "📅 " + now, ""]
    if signals:
        for s in signals:
            icon = "📈" if s["change_pct"] >= 0 else "📉"
            lines += [
                icon + " " + s["ticker"] + "  —  " + str(s["close"]) + " جنيه"
                    + "  (" + ("+{:.2f}".format(s["change_pct"]) if s["change_pct"] >= 0
                               else "{:.2f}".format(s["change_pct"])) + "%)",
                "   RSI: " + str(s["rsi"]) + "  |  ADX: " + str(s["adx"])
                    + "  |  R:R: " + str(s["risk_reward"]) + "x",
                "   🛑 SL: " + str(s["stop_loss"]) + "    🎯 TP: " + str(s["take_profit"]),
                ""
            ]
        lines.append("📊 فُحص " + str(stats["data_ok"]) + " سهم | "
                     + str(stats["no_data"]) + " بدون بيانات")
    else:
        lines += ["لا توجد فرص تحقق الشروط حالياً",
                  "📊 فُحص " + str(stats["data_ok"]) + " سهم | "
                  + str(stats["no_data"]) + " بدون بيانات"]

    text = "\n".join(lines)
    await bot.send_message(chat_id=cfg["telegram_chat_id"], text=text)
    log.info("✅ تليجرام: تم إرسال الرسالة")

    if excel_bytes:
        fname = "EGX_" + datetime.now().strftime("%Y%m%d_%H%M") + ".xlsx"
        if signals:
            caption = "📊 EGX Scanner — " + str(len(signals)) + " فرصة | " + now
        else:
            caption = "📊 EGX Scanner — لا توجد فرص | " + now + " | راجع شيت قريب من الشروط"
        await bot.send_document(
            chat_id=cfg["telegram_chat_id"],
            document=excel_bytes,
            filename=fname,
            caption=caption
        )
        log.info("✅ تليجرام: تم إرسال الـ Excel")

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
async def main():
    log.info("=" * 60)
    log.info(f"   EGX AI Scanner v2  —  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    log.info(f"   عدد الأسهم: {len(EGX_TICKERS)}")
    log.info("=" * 60)

    signals, top3, stats = run_scan(EGX_TICKERS, CONFIG)

    log.info("=" * 60)
    log.info(f"   النتائج: {len(signals)} فرصة | "
             f"قريب من الشروط: {len(top3)} | "
             f"بيانات OK: {stats['data_ok']} | بدون بيانات: {stats['no_data']}")
    log.info("=" * 60)

    if top3:
        log.info("🟡 أقرب الأسهم للشروط:")
        for nm in top3:
            log.info(f"   [{nm['score']}/100] {nm['ticker']}: {nm['missing']}")

    excel = build_excel(signals, stats, near_miss=top3)
    fname = f"EGX_Signals_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    with open(fname, "wb") as f:
        f.write(excel)
    log.info(f"✅ Excel: {fname}")

    await send_telegram(signals, excel, stats, CONFIG, near_miss=top3)


if __name__ == "__main__":
    asyncio.run(main())
